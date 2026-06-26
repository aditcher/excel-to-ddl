"""
Excel → DDL Automator
Business/Data Analyst Portfolio Tool
Author: Aaron Ditcher | github.com/aditcher
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os, re, csv, threading
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── Palette ───────────────────────────────────────────────────────────────────
APP_TITLE   = "Excel → DDL Automator"
APP_VERSION = "1.0.0"

BG         = "#0f1117"
PANEL      = "#1a1f2e"
PANEL2     = "#222840"
BORDER     = "#2e3554"
HEADER_BG  = "#0d1b3e"
TEXT       = "#e8eaf0"
TEXT_MUTED = "#8892b0"
TEXT_DIM   = "#4a5380"
ACCENT     = "#3b82f6"
SUCCESS    = "#22c55e"
WARNING    = "#f59e0b"
DANGER     = "#ef4444"
SQL_CMT    = "#6b7a99"
SQL_KW     = "#60a5fa"
SQL_TYP    = "#4ade80"

MONO   = ("Menlo", 11) if os.name != "nt" else ("Courier New", 10)
SANS   = ("Helvetica", 10)
SANS_B = ("Helvetica", 10, "bold")
SANS_S = ("Helvetica", 9)
SANS_L = ("Helvetica", 11, "bold")

DIALECTS = ["Snowflake","SQL Server (T-SQL)","MySQL","PostgreSQL","Oracle"]
TYPE_MAPS = {
    "Snowflake":          {"INTEGER":"INTEGER","DECIMAL":"FLOAT","TIMESTAMP":"TIMESTAMP_NTZ","DATE":"DATE","BOOLEAN":"BOOLEAN","VARCHAR50":"VARCHAR(50)","VARCHAR255":"VARCHAR(255)","TEXT":"TEXT","VARIANT":"VARIANT"},
    "SQL Server (T-SQL)": {"INTEGER":"INT","DECIMAL":"DECIMAL(18,4)","TIMESTAMP":"DATETIME2","DATE":"DATE","BOOLEAN":"BIT","VARCHAR50":"NVARCHAR(50)","VARCHAR255":"NVARCHAR(255)","TEXT":"NVARCHAR(MAX)","VARIANT":"NVARCHAR(MAX)"},
    "MySQL":              {"INTEGER":"INT","DECIMAL":"DECIMAL(18,4)","TIMESTAMP":"DATETIME","DATE":"DATE","BOOLEAN":"TINYINT(1)","VARCHAR50":"VARCHAR(50)","VARCHAR255":"VARCHAR(255)","TEXT":"TEXT","VARIANT":"JSON"},
    "PostgreSQL":         {"INTEGER":"INTEGER","DECIMAL":"NUMERIC(18,4)","TIMESTAMP":"TIMESTAMP","DATE":"DATE","BOOLEAN":"BOOLEAN","VARCHAR50":"VARCHAR(50)","VARCHAR255":"VARCHAR(255)","TEXT":"TEXT","VARIANT":"JSONB"},
    "Oracle":             {"INTEGER":"NUMBER(10)","DECIMAL":"NUMBER(18,4)","TIMESTAMP":"TIMESTAMP","DATE":"DATE","BOOLEAN":"NUMBER(1)","VARCHAR50":"VARCHAR2(50)","VARCHAR255":"VARCHAR2(255)","TEXT":"CLOB","VARIANT":"CLOB"},
}
DATE_PAT = re.compile(r"^\d{4}-\d{2}-\d{2}$")
DT_PAT   = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}")
BOOL_SET = {"true","false","1","0","yes","no","t","f","y","n"}
FK_PAT   = re.compile(r"(_id|_key|_code|_fk|_ref)$", re.IGNORECASE)

DEMO_HEADERS = ["Customer ID","Full Name","Email Address","Phone #","Date of Birth",
                "Annual Revenue","Is Active","Region Code","Created At","Notes","order_id_ref"]
DEMO_ROWS = [
    ["1001","Alice Johnson","alice@email.com","512-555-0101","1985-03-14","142000.50","true","TX-SW","2023-01-15 09:22:11","VIP customer","5001"],
    ["1002","Bob Smith","bob.smith@corp.net","(512) 555-0202","1990-07-28","88500","1","CA-NO","2023-02-01 00:00:00","Prefers email","5002"],
    ["1003","Carol White","","512.555.0303","1978-11-02","210000.00","true","NY-ME","2023-02-14 14:05:00",None,"5003"],
    ["1004","David Park",None,"512-555-0404","1995-05-30","55000","0","TX-SW",None,"Trial user","5004"],
    ["1005","Eve Torres","eve.t@web.io","","1988-12-19","175500.75","true","FL-SO","2023-03-08 10:11:22","","5005"],
]

# ── Data Cleaning ─────────────────────────────────────────────────────────────

DATE_FMTS = [
    "%Y-%m-%d","%m/%d/%Y","%d/%m/%Y","%m-%d-%Y","%d-%m-%Y",
    "%B %d, %Y","%b %d, %Y","%Y%m%d","%d.%m.%Y","%m.%d.%Y",
]

def _parse_date(s):
    s = str(s).strip()
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return s

def _normalize_company(s):
    if s is None: return None
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    suffixes = [r",?\s*(Inc\.?|LLC\.?|Ltd\.?|Corp\.?|Co\.?|Company|Limited|Incorporated)\.?$"]
    for pat in suffixes:
        s = re.sub(pat, "", s, flags=re.IGNORECASE).strip()
    return s.title()

def clean_rows(headers, rows):
    """
    Returns (cleaned_rows, cleaning_log)
    - Strips whitespace from all cells
    - Normalizes dates to YYYY-MM-DD
    - Normalizes company/name columns (title case, strip suffixes)
    - Converts 'NA', 'N/A', 'n/a', 'NULL', 'None' strings to real NULL
    - Detects zero-as-null in categorical/text columns (size, style, color, type)
    - Removes exact duplicate rows
    - Collapses blank strings to None
    """
    log     = []
    cleaned = []
    seen    = set()

    # detect column roles by header name
    date_cols  = set()
    name_cols  = set()
    cat_cols   = set()   # categorical — zero is not a valid value here

    for i, h in enumerate(headers):
        hl = h.lower()
        if any(x in hl for x in ["date","dob","born","created","updated","modified","timestamp"]):
            date_cols.add(i)
        if any(x in hl for x in ["name","company","vendor","customer","employer","client","organization"]):
            name_cols.add(i)
        if any(x in hl for x in ["color","colour","size","style","type","category","status",
                                   "gender","tier","grade","class","level","flag","code"]):
            cat_cols.add(i)

    # sentinel strings that mean NULL
    NULL_STRINGS = {"na","n/a","null","none","nil","#n/a","#null!","(null)","(none)","unknown",""}

    dupe_count  = 0
    date_fixed  = 0
    name_fixed  = 0
    blank_fixed = 0
    ws_fixed    = 0
    na_fixed    = 0
    zero_fixed  = 0

    # pre-scan: find columns where >50% of non-empty values are "0"
    # those are zero-as-null candidate columns
    zero_null_cols = set()
    for i, h in enumerate(headers):
        vals     = [r[i] if i < len(r) else None for r in rows]
        non_empty = [str(v).strip() for v in vals
                     if v is not None and str(v).strip() != ""]
        if not non_empty: continue
        zero_pct = sum(1 for v in non_empty if v in ("0","0.0")) / len(non_empty)
        # if column is categorical AND >20% zeros, treat zero as null
        if i in cat_cols and zero_pct > 0.2:
            zero_null_cols.add(i)

    for row in rows:
        new_row = []
        for i, v in enumerate(row):
            if v is None:
                new_row.append(None); continue

            s       = str(v)
            stripped = s.strip()

            # whitespace
            if stripped != s:
                ws_fixed += 1

            # blank → NULL
            if stripped == "":
                new_row.append(None); blank_fixed += 1; continue

            # NA-sentinel strings → NULL
            if stripped.lower() in NULL_STRINGS:
                new_row.append(None); na_fixed += 1; continue

            # zero-as-null in categorical columns
            if i in zero_null_cols and stripped in ("0","0.0"):
                new_row.append(None); zero_fixed += 1; continue

            # date normalization
            if i in date_cols:
                normalized = _parse_date(stripped)
                if normalized != stripped:
                    date_fixed += 1; stripped = normalized

            # name/company normalization
            elif i in name_cols:
                normalized = _normalize_company(stripped)
                if normalized != stripped:
                    name_fixed += 1; stripped = normalized

            new_row.append(stripped)

        # deduplicate
        key = tuple(str(x) if x is not None else "" for x in new_row)
        if key in seen:
            dupe_count += 1; continue
        seen.add(key)
        cleaned.append(new_row)

    if dupe_count:  log.append(f"Removed {dupe_count} duplicate row(s)")
    if ws_fixed:    log.append(f"Stripped whitespace from {ws_fixed} cell(s)")
    if blank_fixed: log.append(f"Normalized {blank_fixed} empty string(s) to NULL")
    if na_fixed:    log.append(f"Converted {na_fixed} 'NA'/'N/A'/'NULL' string(s) to NULL")
    if zero_fixed:  log.append(f"Converted {zero_fixed} zero placeholder(s) to NULL in categorical columns")
    if date_fixed:  log.append(f"Standardized {date_fixed} date format(s) to YYYY-MM-DD")
    if name_fixed:  log.append(f"Normalized {name_fixed} name/company value(s)")

    return cleaned, log

# ── Type Inference ─────────────────────────────────────────────────────────────

def slugify(s):
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "col"

def infer_type(values, dialect):
    tmap = TYPE_MAPS.get(dialect, TYPE_MAPS["Snowflake"])
    non  = [v for v in values if v is not None and str(v).strip() != ""]
    if not non: return tmap["TEXT"]
    types_seen = set()
    for v in non:
        s = str(v).strip()
        try:    int(s);   types_seen.add("int");      continue
        except: pass
        try:    float(s); types_seen.add("float");    continue
        except: pass
        if s.lower() in BOOL_SET:   types_seen.add("bool");     continue
        if DT_PAT.match(s):         types_seen.add("datetime"); continue
        if DATE_PAT.match(s):       types_seen.add("date");     continue
        types_seen.add("str")
    if len(types_seen) > 2:                     return tmap["VARIANT"]
    if types_seen <= {"int"}:                   return tmap["INTEGER"]
    if types_seen <= {"int","float"} or types_seen <= {"float"}: return tmap["DECIMAL"]
    if types_seen <= {"bool"}:                  return tmap["BOOLEAN"]
    if types_seen <= {"datetime"}:              return tmap["TIMESTAMP"]
    if types_seen <= {"date"}:                  return tmap["DATE"]
    max_len = max(len(str(v)) for v in non)
    if max_len <= 50:  return tmap["VARCHAR50"]
    if max_len <= 255: return tmap["VARCHAR255"]
    return tmap["TEXT"]

def analyze_columns(headers, rows, dialect):
    cols = []
    total = len(rows)
    for i, h in enumerate(headers):
        vals    = [r[i] if i < len(r) else None for r in rows]
        non     = [v for v in vals if v is not None and str(v).strip() != ""]
        nulls   = total - len(non)
        null_p  = round((nulls / total * 100) if total else 0, 1)
        unique  = len(set(str(v) for v in non))
        is_uniq = unique == len(non) and len(non) > 0
        dtype   = infer_type(vals, dialect)
        is_pk   = i == 0 and is_uniq and dtype in [TYPE_MAPS[dialect].get("INTEGER","INTEGER")]
        is_fk   = bool(FK_PAT.search(h)) and not is_pk
        sample  = ", ".join(str(v) for v in non[:3])
        cols.append({
            "original": h, "name": slugify(h), "type": dtype,
            "null_pct": null_p, "null_count": nulls,
            "unique_count": unique, "total": total,
            "is_pk": is_pk, "is_fk": is_fk,
            "is_unique": is_uniq, "has_null": nulls > 0,
            "has_dupes": not is_uniq and len(non) > 0, "sample": sample,
        })
    return cols

def quote_name(name, dialect):
    if dialect == "MySQL":              return f"`{name}`"
    if dialect == "SQL Server (T-SQL)": return f"[{name}]"
    if dialect == "Oracle":             return f'"{name.upper()}"'
    return f'"{name}"'

def build_ddl(cols, table_name, dialect, null_mode, include_inserts, rows):
    q      = lambda n: quote_name(n, dialect)
    tn     = table_name.upper() if dialect == "Oracle" else table_name
    schema = "PUBLIC." if dialect == "Snowflake" else ""
    ifne   = "" if dialect in ("SQL Server (T-SQL)","Oracle") else " IF NOT EXISTS"
    semi   = "" if dialect == "Oracle" else ";"
    ts     = datetime.now().strftime("%Y-%m-%d %H:%M")
    hdr = (
        f"-- ============================================================\n"
        f"-- Excel -> DDL Automator  |  github.com/aditcher\n"
        f"-- Generated : {ts}\n"
        f"-- Dialect   : {dialect}\n"
        f"-- Table     : {tn}   Columns: {len(cols)}\n"
        f"-- ============================================================\n\n"
    )
    fk_hints = [c for c in cols if c["is_fk"]]
    fk_block = ""
    if fk_hints:
        fk_block = "-- FK CANDIDATES (heuristic - verify before applying)\n"
        for c in fk_hints:
            fk_block += f"--   {q(c['name'])}  ->  possible reference column\n"
        fk_block += "\n"
    col_lines = []
    for c in cols:
        if null_mode == "all_null":  nullable = " NULL"
        elif null_mode == "all_nn":  nullable = " NOT NULL"
        else: nullable = " NOT NULL" if (c["is_pk"] or not c["has_null"]) else " NULL"
        uniq = " UNIQUE" if (not c["is_pk"] and c["is_unique"] and not c["has_null"]) else ""
        col_lines.append(f"    {q(c['name'])} {c['type']}{nullable}{uniq}")
    pk = next((c for c in cols if c["is_pk"]), None)
    if pk:
        col_lines.append(f"    CONSTRAINT pk_{tn} PRIMARY KEY ({q(pk['name'])})")
    create = (f"CREATE TABLE{ifne} {schema}{q(tn)} (\n"
              + ",\n".join(col_lines) + f"\n){semi}\n")
    insert_block = ""
    if include_inserts and rows:
        insert_block = f"\n\n-- SAMPLE INSERTS (first 5 rows)\n"
        col_names = ", ".join(q(c["name"]) for c in cols)
        for row in rows[:5]:
            vals = []
            for i, c in enumerate(cols):
                v = row[i] if i < len(row) else None
                if v is None or str(v).strip() == "":
                    vals.append("NULL")
                else:
                    t = c["type"].upper()
                    if any(x in t for x in ["INT","FLOAT","DECIMAL","NUMERIC","NUMBER","BIT","TINYINT","BOOLEAN"]):
                        try:    float(str(v)); vals.append(str(v))
                        except: vals.append(f"'{str(v)}'")
                    else:
                        vals.append(f"'{str(v).replace(chr(39), chr(39)*2)}'")
            insert_block += f"INSERT INTO {schema}{q(tn)} ({col_names}) VALUES ({', '.join(vals)}){semi}\n"
    return hdr + fk_block + create + insert_block

def build_data_dictionary(cols, table_name):
    return [{
        "Table": table_name, "Column (original)": c["original"],
        "Column (DDL)": c["name"], "Inferred Type": c["type"],
        "Null %": f"{c['null_pct']}%", "Null Count": c["null_count"],
        "Unique Count": c["unique_count"], "Total Rows": c["total"],
        "PK Candidate": "Yes" if c["is_pk"] else "",
        "FK Hint": "Yes" if c["is_fk"] else "",
        "Unique": "Yes" if c["is_unique"] else "",
        "Sample Values": c["sample"],
    } for c in cols]

def save_dict_xlsx(dict_rows, path, table_name, clean_log):
    """Export a formatted Excel data dictionary — the real BA deliverable."""
    if openpyxl is None:
        raise ImportError("openpyxl required")
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data Dictionary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Data Dictionary"

    navy_fill  = PatternFill("solid", fgColor="0D1B3E")
    green_fill = PatternFill("solid", fgColor="1A3A1A")
    alt_fill   = PatternFill("solid", fgColor="1A1F2E")
    white_font = Font(name="Calibri", color="E8EAF0", bold=True, size=10)
    body_font  = Font(name="Calibri", color="E8EAF0", size=10)
    green_font = Font(name="Calibri", color="22C55E", bold=True, size=10)
    warn_font  = Font(name="Calibri", color="F59E0B", bold=True, size=10)
    thin       = Side(style="thin", color="2E3554")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")

    # title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Data Dictionary  —  {table_name}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.fill  = navy_fill
    title_cell.font  = Font(name="Calibri", color="E8EAF0", bold=True, size=12)
    title_cell.alignment = left
    ws.row_dimensions[1].height = 24

    # header row
    headers = ["#","Column (original)","Column (DDL)","Inferred Type",
               "Null %","Null Count","Unique Count","Total Rows",
               "PK","FK Hint","Unique","Sample Values"]
    col_widths = [4, 22, 22, 18, 8, 10, 12, 10, 5, 8, 8, 36]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill      = navy_fill
        cell.font      = white_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 18

    # data rows
    for row_idx, r in enumerate(dict_rows, 3):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="141929")
        vals = [
            row_idx - 2,
            r["Column (original)"], r["Column (DDL)"], r["Inferred Type"],
            r["Null %"], r["Null Count"], r["Unique Count"], r["Total Rows"],
            r["PK Candidate"], r["FK Hint"], r["Unique"], r["Sample Values"],
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.fill   = fill
            cell.border = border
            # color-code special columns
            if col_idx == 9 and v == "Yes":        # PK
                cell.font = green_font
            elif col_idx == 10 and v == "Yes":     # FK
                cell.font = warn_font
            elif col_idx == 5:                     # Null %
                pct = float(str(v).replace("%","") or 0)
                cell.font = Font(name="Calibri", size=10,
                                 color=("EF4444" if pct > 50
                                        else "F59E0B" if pct > 20
                                        else "22C55E" if pct == 0
                                        else "E8EAF0"))
                cell.alignment = center
            else:
                cell.font      = body_font
                cell.alignment = left if col_idx > 1 else center
        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "A3"

    # ── Sheet 2: Cleaning Log ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Cleaning Log")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 60

    ws2.merge_cells("A1:B1")
    c = ws2["A1"]
    c.value     = f"Cleaning Log  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.fill      = navy_fill
    c.font      = Font(name="Calibri", color="E8EAF0", bold=True, size=11)
    c.alignment = left
    ws2.row_dimensions[1].height = 22

    if not clean_log:
        ws2["A2"].value = "No changes — data was already clean."
        ws2["A2"].font  = body_font
    else:
        for i, item in enumerate(clean_log, 2):
            ws2[f"A{i}"].value     = f"  ✓  {item}"
            ws2[f"A{i}"].font      = body_font
            ws2[f"A{i}"].fill      = (alt_fill if i % 2 == 0
                                      else PatternFill("solid", fgColor="141929"))
            ws2[f"A{i}"].alignment = left
            ws2.row_dimensions[i].height = 16

    ws2["A1"].fill = navy_fill

    wb.save(path)

def read_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            data = list(csv.reader(f))
        if not data: return [], [], []
        headers = data[0]
        rows    = [[c if c.strip() != "" else None for c in r]
                   for r in data[1:] if any(c.strip() for c in r)]
        return headers, rows, ["Sheet1"]
    if openpyxl is None:
        raise ImportError("openpyxl is required.\nRun: pip3 install openpyxl")
    wb     = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return None, None, sheets

def read_sheet(path, sheet_name):
    if openpyxl is None:
        raise ImportError("openpyxl not installed")
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb[sheet_name]
    data = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not data: return [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    rows    = [[str(v) if v is not None else None for v in r]
               for r in data[1:] if any(v is not None for v in r)]
    return headers, rows

# ── Dark ttk style ─────────────────────────────────────────────────────────────

def apply_dark_style():
    s = ttk.Style()
    s.theme_use("default")
    s.configure(".", background=PANEL, foreground=TEXT,
                 fieldbackground=PANEL2, troughcolor=PANEL2,
                 selectbackground=ACCENT, selectforeground=TEXT,
                 bordercolor=BORDER, relief="flat", font=SANS)
    s.configure("TNotebook", background=BG, borderwidth=0)
    s.configure("TNotebook.Tab", background=PANEL2, foreground=TEXT_MUTED,
                padding=[14,6], font=SANS, borderwidth=0)
    s.map("TNotebook.Tab",
          background=[("selected", PANEL)],
          foreground=[("selected", TEXT)])
    s.configure("TCombobox", background=PANEL2, foreground=TEXT,
                fieldbackground=PANEL2, arrowcolor=TEXT_MUTED,
                bordercolor=BORDER, selectbackground=PANEL2, selectforeground=TEXT)
    s.map("TCombobox",
          fieldbackground=[("readonly", PANEL2)],
          foreground=[("readonly", TEXT)],
          selectbackground=[("readonly", PANEL2)],
          selectforeground=[("readonly", TEXT)])
    s.configure("Treeview", background=PANEL, foreground=TEXT,
                fieldbackground=PANEL, rowheight=24, borderwidth=0, font=SANS_S)
    s.configure("Treeview.Heading", background=PANEL2, foreground=TEXT_MUTED,
                relief="flat", font=SANS_B, borderwidth=0)
    s.map("Treeview",
          background=[("selected", ACCENT)],
          foreground=[("selected","#ffffff")])
    s.configure("Vertical.TScrollbar", background=PANEL2,
                troughcolor=BG, arrowcolor=PANEL2, borderwidth=0,
                relief="flat", width=8)
    s.configure("Horizontal.TScrollbar", background=PANEL2,
                troughcolor=BG, arrowcolor=PANEL2, borderwidth=0,
                relief="flat", width=8)
    s.map("Vertical.TScrollbar",
          background=[("active", BORDER), ("disabled", BG)])
    s.map("Horizontal.TScrollbar",
          background=[("active", BORDER), ("disabled", BG)])

# ── Label-Button (bypasses macOS Aqua color override) ─────────────────────────

class LBtn(tk.Label):
    def __init__(self, parent, text, cmd, bg=PANEL2, fg=TEXT,
                 font=None, padx=12, pady=5, enabled=True):
        super().__init__(parent, text=text, bg=bg, fg=fg,
                         font=font or SANS, cursor="hand2",
                         padx=padx, pady=pady, relief="flat")
        self._cmd     = cmd
        self._bg      = bg
        self._fg      = fg
        self._enabled = enabled
        if not enabled:
            self.config(fg=TEXT_DIM)
        self.bind("<Enter>",    self._hover_on)
        self.bind("<Leave>",    self._hover_off)
        self.bind("<Button-1>", self._click)

    def _hover_on(self, e):
        if self._enabled:
            try:
                r,g,b = int(self._bg[1:3],16),int(self._bg[3:5],16),int(self._bg[5:7],16)
                h = "#{:02x}{:02x}{:02x}".format(min(255,r+25),min(255,g+25),min(255,b+25))
            except: h = BORDER
            self.config(bg=h)

    def _hover_off(self, e):
        self.config(bg=self._bg)

    def _click(self, e):
        if self._enabled and self._cmd:
            self._cmd()

    def enable(self, bg=None, fg=None):
        self._enabled = True
        if bg: self._bg = bg
        if fg: self._fg = fg
        self.config(bg=self._bg, fg=self._fg)

    def disable(self, bg=None):
        self._enabled = False
        if bg: self._bg = bg
        self.config(bg=self._bg, fg=TEXT_DIM)

    def set_text(self, t):
        self.config(text=t)

# ── App ────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("960x820")
        self.minsize(820, 700)
        self.configure(bg=BG)
        apply_dark_style()
        # grab focus to the window itself so no child widget starts focused
        self.after(100, self.focus_set)
        self._file_path  = None
        self._headers    = []
        self._rows       = []
        self._cols       = []
        self._ddl        = ""
        self._dict_rows  = []
        self._clean_log  = []
        self._cleaned_rows = []
        self._build_ui()

    # ── layout helpers ────────────────────────────────────────────────────────

    def _lbl(self, p, text, fg=TEXT, font=None, bg=PANEL, **kw):
        return tk.Label(p, text=text, bg=bg, fg=fg, font=font or SANS, **kw)

    def _section(self, parent, title):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill="x", padx=10, pady=(8,2))
        tk.Label(f, text=title, bg=BG, fg=ACCENT, font=SANS_B).pack(side="left")
        tk.Frame(f, bg=BORDER, height=1).pack(
            side="left", fill="x", expand=True, padx=(8,0), pady=5)

    def _card(self, parent, expand=False, pady=(0,4)):
        f = tk.Frame(parent, bg=PANEL,
                     highlightthickness=1, highlightbackground=BORDER)
        f.pack(fill="both" if expand else "x",
               expand=expand, padx=10, pady=pady)
        return f

    # ── build ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # header
        hbar = tk.Frame(self, bg=HEADER_BG, height=50)
        hbar.pack(fill="x")
        hbar.pack_propagate(False)
        tk.Label(hbar, text="  Excel → DDL Automator",
                 bg=HEADER_BG, fg=TEXT,
                 font=("Helvetica",13,"bold")).pack(side="left", padx=16)
        tk.Label(hbar, text=f"v{APP_VERSION}   github.com/aditcher",
                 bg=HEADER_BG, fg=TEXT_MUTED, font=SANS_S).pack(side="right", padx=16)

        # top fixed area
        top = tk.Frame(self, bg=BG)
        top.pack(fill="x")

        self._build_upload(top)
        self._build_config(top)
        self._build_audit(top)

        # output fills the rest — no nested scroll
        self._section(self, "4  Output")
        self._build_output(self)

    def _build_upload(self, parent):
        self._section(parent, "1  Upload File")
        card = self._card(parent)
        row  = tk.Frame(card, bg=PANEL)
        row.pack(fill="x", padx=10, pady=10)
        self._file_label = tk.Label(row, text="No file selected",
                                    bg=PANEL, fg=TEXT_MUTED, font=SANS, anchor="w")
        self._file_label.pack(side="left", fill="x", expand=True)
        LBtn(row, "  Browse…  ", self._browse,
             bg=PANEL2, fg=TEXT).pack(side="left", padx=(4,4))
        LBtn(row, "  Load Demo Data  ", self._load_demo,
             bg=ACCENT, fg="#ffffff").pack(side="left", padx=(0,4))

        self._sheet_row = tk.Frame(card, bg=PANEL)
        tk.Label(self._sheet_row, text="Sheet:", bg=PANEL,
                 fg=TEXT_MUTED, font=SANS).pack(side="left", padx=(10,6))
        self._sheet_var = tk.StringVar()
        self._sheet_cb  = ttk.Combobox(self._sheet_row,
                                        textvariable=self._sheet_var,
                                        state="readonly", width=32)
        self._sheet_cb.pack(side="left", pady=(0,8))
        self._sheet_cb.bind("<<ComboboxSelected>>",
                            lambda e: self._load_sheet_threaded(self._sheet_var.get()))

    def _build_config(self, parent):
        self._section(parent, "2  Configure Output")
        card = self._card(parent)
        r1   = tk.Frame(card, bg=PANEL)
        r1.pack(fill="x", padx=10, pady=(10,4))

        tk.Label(r1, text="Table name:", bg=PANEL, fg=TEXT_MUTED,
                 font=SANS).pack(side="left", padx=(0,6))
        self._table_var = tk.StringVar(value="imported_data")
        tk.Entry(r1, textvariable=self._table_var, width=20,
                 bg=PANEL2, fg=TEXT, insertbackground=TEXT, relief="flat",
                 font=SANS, highlightthickness=1,
                 highlightbackground=BORDER,
                 highlightcolor=ACCENT).pack(side="left", padx=(0,16), ipady=4)

        tk.Label(r1, text="Dialect:", bg=PANEL, fg=TEXT_MUTED,
                 font=SANS).pack(side="left", padx=(0,6))
        self._dialect_var = tk.StringVar(value="Snowflake")
        ttk.Combobox(r1, textvariable=self._dialect_var,
                     values=DIALECTS, state="readonly",
                     width=22).pack(side="left", padx=(0,16))

        tk.Label(r1, text="Nullability:", bg=PANEL, fg=TEXT_MUTED,
                 font=SANS).pack(side="left", padx=(0,6))
        self._null_var = tk.StringVar(value="inferred")
        ttk.Combobox(r1, textvariable=self._null_var,
                     values=["inferred","all_null","all_nn"],
                     state="readonly", width=14).pack(side="left")

        r2 = tk.Frame(card, bg=PANEL)
        r2.pack(fill="x", padx=10, pady=(0,4))
        self._inserts_var = tk.BooleanVar(value=False)
        tk.Checkbutton(r2, text="Include sample INSERT statements (first 5 rows)",
                       variable=self._inserts_var,
                       bg=PANEL, fg=TEXT, selectcolor=PANEL2,
                       activebackground=PANEL, activeforeground=TEXT,
                       font=SANS).pack(side="left")
        self._clean_var = tk.BooleanVar(value=True)
        tk.Checkbutton(r2, text="Clean data before generating",
                       variable=self._clean_var,
                       bg=PANEL, fg=TEXT, selectcolor=PANEL2,
                       activebackground=PANEL, activeforeground=TEXT,
                       font=SANS).pack(side="left", padx=(20,0))

        r3 = tk.Frame(card, bg=PANEL)
        r3.pack(pady=(2,10))
        self._gen_btn = LBtn(r3,
                             "  ⚡ Generate DDL + Data Dictionary  ",
                             self._generate,
                             bg="#1a3a1a", fg=TEXT_DIM,
                             font=SANS_L, padx=20, pady=8,
                             enabled=False)
        self._gen_btn.pack()

    def _build_audit(self, parent):
        self._section(parent, "3  Data Quality Audit")
        self._audit_card = self._card(parent)
        tk.Label(self._audit_card,
                 text="Load a file or demo data to see the audit.",
                 bg=PANEL, fg=TEXT_MUTED, font=SANS).pack(pady=12)

    def _build_output(self, parent):
        # output card expands to fill remaining window height
        card = tk.Frame(parent, bg=PANEL,
                        highlightthickness=1, highlightbackground=BORDER)
        card.pack(fill="both", expand=True, padx=10, pady=(0,10))

        # toolbar row
        tb = tk.Frame(card, bg=PANEL2, height=38)
        tb.pack(fill="x")
        tb.pack_propagate(False)
        for label, cmd in [("  Copy DDL  ", self._copy_ddl),
                            ("  Save .sql  ", self._save_sql),
                            ("  Save Dict .csv  ", self._save_dict),
                            ("  Save Dict .xlsx  ", self._save_dict_xlsx)]:
            LBtn(tb, label, cmd, bg=PANEL2, fg=TEXT,
                 padx=10, pady=6).pack(side="left", padx=2, pady=3)

        # notebook tabs — fills remaining card space
        nb = ttk.Notebook(card)
        nb.pack(fill="both", expand=True)

        # DDL tab
        ddl_f = tk.Frame(nb, bg="#0d1117")
        nb.add(ddl_f, text="  DDL Output  ")
        self._ddl_text = tk.Text(
            ddl_f, font=MONO, bg="#0d1117", fg="#e6edf3",
            insertbackground="#0d1117", relief="flat", wrap="word",
            padx=14, pady=12,
            spacing1=0, spacing2=0, spacing3=0,
            highlightthickness=0, borderwidth=0,
            highlightbackground="#0d1117", highlightcolor="#0d1117",
            takefocus=0,
            selectbackground=ACCENT, selectforeground="#ffffff")
        # steal focus back to the window immediately on any click
        # so macOS never draws its Aqua focus ring on the Text widget
        self._ddl_text.bind("<Button-1>", lambda e: self.after(1, lambda: self.focus_set()))
        self._ddl_text.bind("<FocusIn>",  lambda e: self.after(1, lambda: self.focus_set()))
        sy = ttk.Scrollbar(ddl_f, orient="vertical", command=self._ddl_text.yview)
        self._ddl_text.configure(yscrollcommand=sy.set)
        sy.pack(side="right", fill="y")
        self._ddl_text.pack(fill="both", expand=True)
        self._ddl_text.tag_configure("kw",  foreground=SQL_KW)
        self._ddl_text.tag_configure("typ", foreground=SQL_TYP)
        self._ddl_text.tag_configure("cmt", foreground=SQL_CMT)
        self._set_ddl(f"-- {APP_TITLE}\n-- Load a file and click Generate to see output.\n")

        # Data Dictionary tab
        dict_f = tk.Frame(nb, bg=PANEL)
        nb.add(dict_f, text="  Data Dictionary  ")
        dcols  = ["Column (original)","Column (DDL)","Inferred Type",
                  "Null %","Null Count","Unique Count","PK","FK","Sample Values"]
        self._dict_tree = ttk.Treeview(dict_f, columns=dcols, show="headings")
        for col, w in zip(dcols,[140,120,110,60,80,90,40,40,220]):
            self._dict_tree.heading(col, text=col)
            self._dict_tree.column(col, width=w, minwidth=40)
        dsy = ttk.Scrollbar(dict_f, orient="vertical",   command=self._dict_tree.yview)
        dsx = ttk.Scrollbar(dict_f, orient="horizontal", command=self._dict_tree.xview)
        self._dict_tree.configure(yscrollcommand=dsy.set, xscrollcommand=dsx.set)
        dsy.pack(side="right", fill="y")
        dsx.pack(side="bottom", fill="x")
        self._dict_tree.pack(fill="both", expand=True)

        # Column Name Map tab
        san_f = tk.Frame(nb, bg=PANEL)
        nb.add(san_f, text="  Column Name Map  ")
        scols = ["Original Name","DDL Name","Changed?"]
        self._san_tree = ttk.Treeview(san_f, columns=scols, show="headings")
        for col in scols:
            self._san_tree.heading(col, text=col)
            self._san_tree.column(col, width=240, minwidth=80)
        ssy = ttk.Scrollbar(san_f, orient="vertical", command=self._san_tree.yview)
        self._san_tree.configure(yscrollcommand=ssy.set)
        ssy.pack(side="right", fill="y")
        self._san_tree.pack(fill="both", expand=True)
        self._san_tree.tag_configure("changed", foreground=WARNING)

        # Cleaning Log tab
        log_f = tk.Frame(nb, bg=PANEL)
        nb.add(log_f, text="  Cleaning Log  ")
        self._log_text = tk.Text(
            log_f, font=SANS, bg=PANEL, fg=TEXT,
            relief="flat", padx=14, pady=12, wrap="word",
            highlightthickness=0, borderwidth=0,
            highlightbackground=PANEL, highlightcolor=PANEL,
            takefocus=0, state="disabled")
        self._log_text.bind("<Button-1>", lambda e: self.after(1, lambda: self.focus_set()))
        self._log_text.bind("<FocusIn>",  lambda e: self.after(1, lambda: self.focus_set()))
        lsy = ttk.Scrollbar(log_f, orient="vertical", command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=lsy.set)
        lsy.pack(side="right", fill="y")
        self._log_text.pack(fill="both", expand=True)

        # Row Preview tab
        self._prev_f = tk.Frame(nb, bg=PANEL)
        nb.add(self._prev_f, text="  Row Preview  ")
        self._prev_label = tk.Label(self._prev_f,
            text="Generate DDL to see a preview of cleaned rows.",
            bg=PANEL, fg=TEXT_MUTED, font=SANS)
        self._prev_label.pack(pady=12)

    # ── data actions ──────────────────────────────────────────────────────────

    def _load_demo(self):
        self._file_path = None
        self._headers   = DEMO_HEADERS[:]
        self._rows      = [r[:] for r in DEMO_ROWS]
        self._sheet_row.pack_forget()
        self._file_label.config(text="Demo data loaded  —  5 rows", fg=SUCCESS)
        self._refresh_audit()
        self._gen_btn.enable(bg=SUCCESS, fg="#ffffff")

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select spreadsheet",
            filetypes=[("Spreadsheets","*.xlsx *.xls *.csv"),("All","*.*")])
        if not path: return
        self._file_path = path
        fname = os.path.basename(path)
        self._file_label.config(text=f"Reading {fname}…", fg=WARNING)
        self._gen_btn.disable(bg="#1a3a1a")
        self.update_idletasks()

        def _do():
            try:
                headers, rows, sheets = read_file(path)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Read error", str(e)))
                self.after(0, lambda: self._file_label.config(
                    text="Error reading file", fg=DANGER))
                return
            def _done():
                self._sheets = sheets
                if len(sheets) > 1:
                    self._sheet_cb["values"] = sheets
                    self._sheet_var.set(sheets[0])
                    self._sheet_row.pack(fill="x", padx=4, pady=(0,6))
                    self._file_label.config(text=fname, fg=TEXT)
                    self._load_sheet_threaded(sheets[0])
                else:
                    self._sheet_row.pack_forget()
                    if headers is not None:
                        self._headers = headers
                        self._rows    = rows
                        self._file_label.config(
                            text=f"{fname}  —  {len(rows):,} rows", fg=TEXT)
                        self._refresh_audit()
                        self._gen_btn.enable(bg=SUCCESS, fg="#ffffff")
            self.after(0, _done)
        threading.Thread(target=_do, daemon=True).start()

    def _load_sheet_threaded(self, sheet):
        self._file_label.config(
            text=f"Loading sheet: {sheet}…", fg=WARNING)
        self._gen_btn.disable(bg="#1a3a1a")
        self.update_idletasks()

        def _do():
            try:
                headers, rows = read_sheet(self._file_path, sheet)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Sheet error", str(e)))
                return
            def _done():
                self._headers = headers
                self._rows    = rows
                fname = os.path.basename(self._file_path)
                self._file_label.config(
                    text=f"{fname}  —  {len(rows):,} rows  [sheet: {sheet}]",
                    fg=TEXT)
                self._refresh_audit()
                self._gen_btn.enable(bg=SUCCESS, fg="#ffffff")
            self.after(0, _done)
        threading.Thread(target=_do, daemon=True).start()

    def _refresh_audit(self):
        dialect    = self._dialect_var.get()
        self._cols = analyze_columns(self._headers, self._rows, dialect)
        for w in self._audit_card.winfo_children():
            w.destroy()

        total       = len(self._rows)
        null_total  = sum(c["null_count"]  for c in self._cols)
        dupe_pk     = any(c["is_pk"] and c["has_dupes"] for c in self._cols)
        fk_count    = sum(1 for c in self._cols if c["is_fk"])
        mixed_count = sum(1 for c in self._cols
                          if any(x in c["type"]
                                 for x in ["VARIANT","JSON","JSONB","CLOB"]))

        stats = [
            ("Rows",       f"{total:,}",  TEXT),
            ("Columns",    str(len(self._cols)), TEXT),
            ("Null cells", str(null_total),
             DANGER if null_total else SUCCESS),
            ("Dupe PK",    "Yes" if dupe_pk else "No",
             DANGER if dupe_pk else SUCCESS),
            ("FK hints",   str(fk_count),
             WARNING if fk_count else TEXT_MUTED),
            ("Mixed types",str(mixed_count),
             WARNING if mixed_count else TEXT_MUTED),
        ]
        stat_row = tk.Frame(self._audit_card, bg=PANEL)
        stat_row.pack(fill="x", padx=10, pady=8)
        for label, val, color in stats:
            box = tk.Frame(stat_row, bg=PANEL2,
                           highlightthickness=1, highlightbackground=BORDER)
            box.pack(side="left", padx=4, ipadx=10, ipady=6)
            tk.Label(box, text=val, bg=PANEL2, fg=color,
                     font=("Helvetica",15,"bold")).pack()
            tk.Label(box, text=label, bg=PANEL2, fg=TEXT_MUTED,
                     font=SANS_S).pack()

        # null bar
        tk.Label(self._audit_card, text="Null % per column:",
                 bg=PANEL, fg=TEXT_MUTED, font=SANS_S
                 ).pack(anchor="w", padx=14, pady=(2,0))
        self._bar = tk.Canvas(self._audit_card, bg=PANEL,
                               height=32, highlightthickness=0)
        self._bar.pack(fill="x", padx=10, pady=(2,8))

        def draw(e=None):
            c = self._bar; c.delete("all")
            w = c.winfo_width()
            if w <= 1 or not self._cols: return
            n, gap = len(self._cols), 3
            bw = max(5, (w - gap*(n+1)) / n)
            for i, col in enumerate(self._cols):
                x0 = gap + i*(bw+gap); x1 = x0+bw
                c.create_rectangle(x0, 4, x1, 26, fill=PANEL2, outline="")
                if col["null_pct"] > 0:
                    fh = int(22 * col["null_pct"] / 100)
                    color = (DANGER if col["null_pct"]>50
                             else WARNING if col["null_pct"]>20 else ACCENT)
                    c.create_rectangle(x0, 26-fh, x1, 26, fill=color, outline="")
        self._bar.bind("<Configure>", draw)
        self.after(80, draw)

    def _generate(self):
        if not self._headers:
            messagebox.showwarning("No data","Load a file or demo data first.")
            return
        dialect    = self._dialect_var.get()
        table_name = slugify(self._table_var.get()) or "imported_data"
        null_mode  = self._null_var.get()
        inserts    = self._inserts_var.get()
        do_clean   = self._clean_var.get()

        self._gen_btn.disable(bg="#1a3a1a")
        self._gen_btn.set_text("  Generating…  ")
        self.update_idletasks()

        def _do():
            rows = self._rows
            log  = []
            if do_clean:
                rows, log = clean_rows(self._headers, rows)
            cols      = analyze_columns(self._headers, rows, dialect)
            ddl       = build_ddl(cols, table_name, dialect,
                                  null_mode, inserts, rows)
            dict_rows = build_data_dictionary(cols, table_name)

            def _done():
                self._cols         = cols
                self._ddl          = ddl
                self._dict_rows    = dict_rows
                self._clean_log    = log
                self._cleaned_rows = rows
                self._set_ddl(ddl)
                self._populate_dict()
                self._populate_san()
                self._populate_log(log, len(self._rows), len(rows))
                self._populate_preview(rows)
                self._gen_btn.enable(bg=SUCCESS, fg="#ffffff")
                self._gen_btn.set_text("  ⚡ Generate DDL + Data Dictionary  ")
            self.after(0, _done)
        threading.Thread(target=_do, daemon=True).start()

    # ── output helpers ────────────────────────────────────────────────────────

    def _set_ddl(self, text):
        self._ddl_text.config(state="normal")
        self._ddl_text.delete("1.0","end")
        self._ddl_text.insert("1.0", text)
        self._ddl_text.tag_remove("kw",  "1.0","end")
        self._ddl_text.tag_remove("cmt", "1.0","end")
        KWS = ["CREATE","TABLE","PRIMARY","KEY","CONSTRAINT","NOT","NULL",
               "IF","EXISTS","UNIQUE","INSERT","INTO","VALUES"]
        for i, line in enumerate(text.split("\n"), 1):
            if line.strip().startswith("--"):
                self._ddl_text.tag_add("cmt", f"{i}.0", f"{i}.end")
            else:
                for kw in KWS:
                    start = f"{i}.0"
                    while True:
                        pos = self._ddl_text.search(
                            r'\b'+kw+r'\b', start,
                            stopindex=f"{i+1}.0", regexp=True)
                        if not pos: break
                        self._ddl_text.tag_add("kw", pos, f"{pos}+{len(kw)}c")
                        start = f"{pos}+{len(kw)}c"
        self._ddl_text.config(state="disabled")

    def _populate_dict(self):
        self._dict_tree.delete(*self._dict_tree.get_children())
        for r in self._dict_rows:
            self._dict_tree.insert("","end", values=(
                r["Column (original)"], r["Column (DDL)"], r["Inferred Type"],
                r["Null %"], r["Null Count"], r["Unique Count"],
                r["PK Candidate"], r["FK Hint"], r["Sample Values"]))

    def _populate_san(self):
        self._san_tree.delete(*self._san_tree.get_children())
        for c in self._cols:
            changed = "Yes  ✎" if c["original"] != c["name"] else ""
            self._san_tree.insert("","end",
                values=(c["original"], c["name"], changed),
                tags=("changed",) if changed else ())

    def _populate_log(self, log, before, after):
        self._log_text.config(state="normal")
        self._log_text.delete("1.0","end")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._log_text.insert("end", f"Cleaning report  —  {ts}\n\n")
        if not log:
            self._log_text.insert("end",
                "No changes made. Data was already clean.\n")
        else:
            removed = before - after
            self._log_text.insert("end",
                f"Rows before: {before:,}   →   Rows after: {after:,}"
                + (f"   ({removed:,} removed)\n\n" if removed else "\n\n"))
            for item in log:
                self._log_text.insert("end", f"  ✓  {item}\n")
        self._log_text.config(state="disabled")

    def _copy_ddl(self):
        if not self._ddl:
            messagebox.showinfo("Nothing to copy","Generate DDL first."); return
        self.clipboard_clear(); self.clipboard_append(self._ddl)
        messagebox.showinfo("Copied","DDL copied to clipboard.")

    def _save_sql(self):
        if not self._ddl:
            messagebox.showinfo("Nothing to save","Generate DDL first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".sql",
            initialfile=f"{self._table_var.get()}_ddl.sql",
            filetypes=[("SQL","*.sql"),("All","*.*")])
        if path:
            with open(path,"w",encoding="utf-8") as f: f.write(self._ddl)
            messagebox.showinfo("Saved", f"Saved to:\n{path}")

    def _populate_preview(self, rows):
        """Build a live scrollable grid of the first 50 cleaned rows."""
        # clear previous content
        for w in self._prev_f.winfo_children():
            w.destroy()

        if not rows:
            tk.Label(self._prev_f, text="No rows to preview.",
                     bg=PANEL, fg=TEXT_MUTED, font=SANS).pack(pady=12)
            return

        preview_rows = rows[:50]
        total        = len(rows)
        shown        = len(preview_rows)

        # info bar
        info = tk.Frame(self._prev_f, bg=PANEL2, height=28)
        info.pack(fill="x")
        info.pack_propagate(False)
        tk.Label(info,
                 text=f"  Showing {shown:,} of {total:,} cleaned rows"
                      + ("  (first 50 shown)" if total > 50 else ""),
                 bg=PANEL2, fg=TEXT_MUTED, font=SANS_S).pack(side="left", padx=8)

        # treeview
        ddl_names = [c["name"] for c in self._cols]
        tree = ttk.Treeview(self._prev_f, columns=ddl_names,
                            show="headings")
        for name in ddl_names:
            tree.heading(name, text=name)
            tree.column(name, width=120, minwidth=60, stretch=True)

        psy = ttk.Scrollbar(self._prev_f, orient="vertical",   command=tree.yview)
        psx = ttk.Scrollbar(self._prev_f, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=psy.set, xscrollcommand=psx.set)
        psy.pack(side="right",  fill="y")
        psx.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        # tag alternating rows
        tree.tag_configure("odd",  background=PANEL)
        tree.tag_configure("even", background=PANEL2)
        tree.tag_configure("null", foreground=TEXT_DIM)

        for i, row in enumerate(preview_rows):
            tag   = "odd" if i % 2 == 0 else "even"
            vals  = []
            has_null = False
            for j in range(len(ddl_names)):
                v = row[j] if j < len(row) else None
                if v is None:
                    vals.append("NULL"); has_null = True
                else:
                    vals.append(str(v))
            tree.insert("", "end", values=vals,
                        tags=(tag, "null") if has_null else (tag,))

    def _save_dict_xlsx(self):
        if not self._dict_rows:
            messagebox.showinfo("Nothing to save","Generate DDL first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{self._table_var.get()}_data_dictionary.xlsx",
            filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if not path: return
        try:
            save_dict_xlsx(self._dict_rows, path,
                           self._table_var.get(),
                           getattr(self, "_clean_log", []))
            messagebox.showinfo("Saved", f"Excel dictionary saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def _save_dict(self):
        if not self._dict_rows:
            messagebox.showinfo("Nothing to save","Generate DDL first."); return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=f"{self._table_var.get()}_data_dictionary.csv",
            filetypes=[("CSV","*.csv"),("All","*.*")])
        if path:
            with open(path,"w",newline="",encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=self._dict_rows[0].keys())
                w.writeheader(); w.writerows(self._dict_rows)
            messagebox.showinfo("Saved", f"Saved to:\n{path}")

if __name__ == "__main__":
    App().mainloop()
